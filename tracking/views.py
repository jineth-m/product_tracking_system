import csv
from openpyxl import Workbook
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils.timezone import localtime
from django.utils import timezone


from .models import (
    Product,
    ProductSubPart,
    SubPart,
    Status,
    SubPartStatusHistory,
    ProductMainType,
    ProductRange,
)
from accounts.models import UserProfile

# ----------------------------
# PRODUCT LIST (Grouped by Type → Range)
# ----------------------------
def product_list(request):
    query = request.GET.get("q", "").strip()

    # Prefetch hierarchy
    main_types = ProductMainType.objects.prefetch_related(
        "ranges__products"
    )

    grouped = []

    for main in main_types:
        range_blocks = []

        for prange in main.ranges.all():

            products = prange.products.all()

            if query:
                products = products.filter(
                    Q(product_code__icontains=query) |
                    Q(product_name__icontains=query)
                )

            product_data = []

            for product in products:
                total = ProductSubPart.objects.filter(product=product).count()
                completed = ProductSubPart.objects.filter(
                    product=product,
                    current_status__name="Completed"
                ).count()

                progress = int((completed / total) * 100) if total else 0

                product_data.append({
                    "product": product,
                    "progress": progress,
                })

            range_blocks.append({
                "range": prange,
                "products": product_data,
            })

        grouped.append({
            "main": main,
            "ranges": range_blocks,
        })

    return render(request, "tracking/product_list.html", {
        "grouped": grouped,
        "query": query,
    })


# ----------------------------
# PRODUCT DETAIL
# ----------------------------
def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    subparts = ProductSubPart.objects.filter(product=product).select_related(
        'sub_part',
        'current_status',
        'updated_by'
    )

    total = subparts.count()
    completed = subparts.filter(current_status__name='Completed').count()
    progress = int((completed / total) * 100) if total > 0 else 0

    return render(request, 'tracking/product_detail.html', {
        'product': product,
        'subparts': subparts,
        'progress': progress,
    })


# ----------------------------
# SUB PART HISTORY
# ----------------------------
def subpart_history(request, subpart_id):
    subpart = get_object_or_404(SubPart, id=subpart_id)

    history = SubPartStatusHistory.objects.filter(
        sub_part=subpart
    ).select_related(
        'old_status',
        'new_status',
        'changed_by'
    ).order_by('-changed_at')

    return render(
        request,
        'tracking/subpart_history.html',
        {
            'subpart': subpart,
            'history': history,
        }
    )



# ----------------------------
# PIC DASHBOARD (FIXED)
# ----------------------------
@login_required
def pic_dashboard(request):
    query = request.GET.get("q", "").strip()

    # --------------------------------------------------
    # 1. Determine visible sub parts
    # --------------------------------------------------
    if request.user.is_superuser:
        sub_parts = SubPart.objects.all()
    else:
        user_profile = request.user.userprofile
        sub_parts = SubPart.objects.filter(
            sub_part_type=user_profile.sub_part_type
        )

    # --------------------------------------------------
    # 2. Search
    # --------------------------------------------------
    if query:
        sub_parts = sub_parts.filter(
            Q(sub_part_code__icontains=query) |
            Q(sub_part_name__icontains=query)
        )

    # --------------------------------------------------
    # 3. Handle SAVE (POST)
    # --------------------------------------------------
    if request.method == "POST":
        sub_part_id = request.POST.get("sub_part_id")
        status_id = request.POST.get("status")
        comment = request.POST.get("comment", "").strip()

        if sub_part_id and status_id:
            sub_part = SubPart.objects.get(id=sub_part_id)
            new_status = Status.objects.get(id=status_id)

            product_subparts = list(
                ProductSubPart.objects.filter(sub_part_id=sub_part_id)
            )

            # -----------------------------------------
            # CASE 1: linked to products
            # -----------------------------------------
            if product_subparts:
                representative = product_subparts[0]
                old_status = representative.current_status

                for psp in product_subparts:
                    psp.current_status = new_status
                    psp.comment = comment
                    psp.updated_by = request.user
                    psp.save()

            # -----------------------------------------
            # CASE 2: NOT linked to any product
            # -----------------------------------------
            else:
                old_status = sub_part.default_status

                sub_part.default_status = new_status
                sub_part.default_comment = comment
                sub_part.default_updated_by = request.user
                sub_part.default_updated_at = timezone.now()
                sub_part.save()

            # 🔴 SAFETY: history cannot accept NULL
            if old_status is None:
                old_status = new_status

            SubPartStatusHistory.objects.create(
                sub_part=sub_part,
                old_status=old_status,
                new_status=new_status,
                comment=comment,
                changed_by=request.user,
            )

        return redirect("pic_dashboard")


    # --------------------------------------------------
    # 4. Build dashboard rows (SHOW ALL SUB PARTS)
    # --------------------------------------------------
    rows = []

    for sub_part in sub_parts:
        links = list(
            ProductSubPart.objects.filter(sub_part=sub_part)
            .select_related("product", "current_status")
        )

        if links:
            representative = links[0]

            rows.append({
                "sub_part": sub_part,
                "sub_part_type": sub_part.sub_part_type,
                "products": [l.product for l in links],
                "current_status": representative.current_status,
                "comment": representative.comment,
                "updated_at": representative.updated_at,
            })

    
        else:
            rows.append({
                "sub_part": sub_part,
                "products": [],
                "current_status": sub_part.default_status,
                "comment": sub_part.default_comment,
                "updated_at": sub_part.default_updated_at,
            })


    # --------------------------------------------------
    # 5. Render
    # --------------------------------------------------
    return render(request, "tracking/pic_dashboard.html", {
        "rows": rows,
        "statuses": Status.objects.all(),
        "query": query,
    })


# ----------------------------
# REPORTS
# ----------------------------
@login_required
def reports(request):
    products = Product.objects.all()
    data = []

    for product in products:
        subparts = ProductSubPart.objects.filter(product=product)

        total = subparts.count()
        completed = subparts.filter(
            current_status__name__iexact="Completed"
        ).count()

        pending = total - completed
        progress = int((completed / total) * 100) if total > 0 else 0

        data.append({
            "product": product,
            "total": total,
            "completed": completed,
            "pending": pending,
            "progress": progress,
        })

    return render(
        request,
        "tracking/reports_product_summary.html",
        {"data": data}
    )

# ----------------------------
# EXPORTS
# ----------------------------
@login_required
def export_product_summary_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product_summary.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Product Code',
        'Product Name',
        'Total Sub Parts',
        'Completed',
        'Pending',
        'Progress (%)'
    ])

    for product in Product.objects.all():
        total = ProductSubPart.objects.filter(product=product).count()
        completed = ProductSubPart.objects.filter(
            product=product,
            current_status__name='Completed'
        ).count()
        progress = int((completed / total) * 100) if total else 0

        writer.writerow([
            product.product_code,
            product.product_name,
            total,
            completed,
            total - completed,
            progress
        ])

    return response


@login_required
def export_product_summary_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Summary"

    ws.append([
        'Product Code',
        'Product Name',
        'Total Sub Parts',
        'Completed',
        'Pending',
        'Progress (%)'
    ])

    for product in Product.objects.all():
        total = ProductSubPart.objects.filter(product=product).count()
        completed = ProductSubPart.objects.filter(
            product=product,
            current_status__name='Completed'
        ).count()
        progress = int((completed / total) * 100) if total else 0

        ws.append([
            product.product_code,
            product.product_name,
            total,
            completed,
            total - completed,
            progress
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="product_summary.xlsx"'
    wb.save(response)

    return response


@login_required
def export_status_history_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="status_history.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Sub Part Code',
        'Product Code',
        'Old Status',
        'New Status',
        'Comment',
        'Changed By',
        'Changed At'
    ])

    history = SubPartStatusHistory.objects.select_related(
        'sub_part', 'product', 'old_status', 'new_status', 'changed_by'
    ).order_by('-changed_at')

    for h in history:
        writer.writerow([
            h.sub_part.sub_part_code,
            h.product.product_code,
            h.old_status.name,
            h.new_status.name,
            h.comment,
            h.changed_by.username if h.changed_by else '',
            h.changed_at
        ])

    return response


@login_required
def export_status_history_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Status History"

    ws.append([
        'Sub Part Code',
        'Product Code',
        'Old Status',
        'New Status',
        'Comment',
        'Changed By',
        'Changed At'
    ])

    history = SubPartStatusHistory.objects.select_related(
        'sub_part', 'product', 'old_status', 'new_status', 'changed_by'
    ).order_by('-changed_at')

    for h in history:
        ws.append([
            h.sub_part.sub_part_code,
            h.product.product_code,
            h.old_status.name,
            h.new_status.name,
            h.comment,
            h.changed_by.username if h.changed_by else '',
            h.changed_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="status_history.xlsx"'
    wb.save(response)

    return response


def product_progress_report(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    subparts = ProductSubPart.objects.filter(
        product=product
    ).select_related(
        "sub_part", "current_status", "updated_by"
    ).order_by("sub_part__sub_part_code")

    return render(request, "tracking/product_progress.html", {
        "product": product,
        "subparts": subparts,
    })

def product_progress_excel(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    subparts = ProductSubPart.objects.filter(
        product=product
    ).select_related("sub_part", "current_status", "updated_by")

    wb = Workbook()
    ws = wb.active
    ws.title = "Progress"

    ws.append([
        "Sub Part Code",
        "Sub Part Name",
        "Status",
        "Comment",
        "Updated By",
        "Last Updated",
    ])

    for psp in subparts:
        ws.append([
            psp.sub_part.sub_part_code,
            psp.sub_part.sub_part_name,
            psp.current_status.name,
            psp.comment,
            psp.updated_by.username if psp.updated_by else "",
            psp.updated_at.strftime("%Y-%m-%d %H:%M"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{product.product_code}_progress.xlsx"'
    )

    wb.save(response)
    return response


def product_list_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Progress"

    # Header row
    ws.append([
        "Product Code",
        "Product Name",
        "Total Sub Parts",
        "Completed",
        "Pending",
        "Progress (%)"
    ])

    products = Product.objects.all()

    for product in products:
        total = ProductSubPart.objects.filter(product=product).count()
        completed = ProductSubPart.objects.filter(
            product=product,
            current_status__name="Completed"
        ).count()
        pending = total - completed
        progress = int((completed / total) * 100) if total else 0

        ws.append([
            product.product_code,
            product.product_name,
            total,
            completed,
            pending,
            progress
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="product_progress_summary.xlsx"'
    )

    wb.save(response)
    return response




@login_required
def export_pic_dashboard_excel(request):

    # same logic as dashboard — reuse it
    if request.user.is_superuser:
        sub_parts = SubPart.objects.all()
    else:
        user_profile = request.user.userprofile
        sub_parts = SubPart.objects.filter(
            sub_part_type=user_profile.sub_part_type
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "PIC Dashboard"

    # Header row
    ws.append([
        "Sub Part Code",
        "Sub Part Name",
        "Sub Part Type",
        "Products",
        "Status",
        "Comment",
        "Updated By",
        "Updated At",
    ])

    for sub_part in sub_parts:
        links = list(
            ProductSubPart.objects.filter(sub_part=sub_part)
            .select_related("product", "current_status", "updated_by")
        )

        if links:
            rep = links[0]

            product_names = ", ".join(
                [l.product.product_code for l in links]
            )

            status = rep.current_status.name if rep.current_status else ""
            comment = rep.comment or ""
            updated_by = rep.updated_by.username if rep.updated_by else ""
            updated_at = localtime(rep.updated_at).strftime("%Y-%m-%d %H:%M")

        else:
            product_names = ""
            status = ""
            comment = ""
            updated_by = ""
            updated_at = ""

        ws.append([
            sub_part.sub_part_code,
            sub_part.sub_part_name,
            sub_part.sub_part_type.name,
            product_names,
            status,
            comment,
            updated_by,
            updated_at,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pic_dashboard.xlsx"'

    wb.save(response)
    return response


def export_range_excel(request, range_id):
    prange = get_object_or_404(ProductRange, id=range_id)
    products = prange.products.all()

    wb = Workbook()
    ws = wb.active
    ws.title = prange.name

    ws.append([
        "Product Code",
        "Product Name",
        "Progress (%)"
    ])

    for product in products:
        total = ProductSubPart.objects.filter(product=product).count()
        completed = ProductSubPart.objects.filter(
            product=product,
            current_status__name="Completed"
        ).count()

        progress = int((completed / total) * 100) if total else 0

        ws.append([
            product.product_code,
            product.product_name,
            progress
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{prange.name}_products.xlsx"'
    )

    wb.save(response)
    return response
